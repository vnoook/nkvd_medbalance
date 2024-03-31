# TODO
# взять файл
# определить его расширение
# zip распаковать и взять из него csv, а csv сразу пустить в работу
# взять инфу в список из нужных колонок
# сделать новые списки относительно алгоритма подсчёта остатков
# создать файл xls и сделать в нём два листа, для подсчёта общего количества и детального списка
#
# Остаток поштучно-каждая строка -1 шт
# Общее количество через фильтр по наименованию, потом по серии и количеству sgtin (или количество строчек)

import os
import sys
import PyQt5
import PyQt5.QtWidgets
import PyQt5.QtCore
import PyQt5.QtGui
import csv
import zipfile
import difflib
import chardet
import tempfile
import pandas as pd
# import locale
import openpyxl
# import openpyxl.utils
# import openpyxl.styles
# import datetime
# import shutil
# import psutil


# класс главного окна
class WindowMain(PyQt5.QtWidgets.QMainWindow):
    """Класс главного окна"""

    # описание главного окна
    def __init__(self):
        super().__init__()

        # главное окно, надпись на нём и размеры
        self.setWindowTitle('Помощник проверки остатков')
        self.setGeometry(150, 150, 1000, 400)
        # self.setWindowFlags(PyQt5.QtCore.Qt.WindowStaysOnTopHint)

        # переменные
        self.info_extention_open_file = 'Файлы (*.zip; *.csv)'
        self.info_path_open_file = None
        self.info_for_open_file = 'Выберите файл (.ZIP) или (.CSV)'
        self.text_empty_path_file = 'файл пока не выбран'
        self.selected_file = None
        self.headers = ['prod_name', 'full_prod_name', 'status', 'sgtin']

        # ОБЪЕКТЫ НА ФОРМЕ
        # label_prompt_select_file
        self.label_prompt_select_file = PyQt5.QtWidgets.QLabel(self)
        self.label_prompt_select_file.setObjectName('label_prompt_select_file')
        self.label_prompt_select_file.setText('Выберите файл ZIP или CSV')
        self.label_prompt_select_file.setGeometry(PyQt5.QtCore.QRect(10, 10, 150, 40))
        font = PyQt5.QtGui.QFont()
        font.setPointSize(12)
        self.label_prompt_select_file.setFont(font)
        self.label_prompt_select_file.adjustSize()
        self.label_prompt_select_file.setToolTip(self.label_prompt_select_file.objectName())

        # button_select_file
        self.button_select_file = PyQt5.QtWidgets.QPushButton(self)
        self.button_select_file.setObjectName('button_select_file')
        self.button_select_file.setText('...')
        self.button_select_file.setGeometry(PyQt5.QtCore.QRect(10, 40, 50, 20))
        self.button_select_file.setFixedWidth(50)
        self.button_select_file.clicked.connect(self.select_file)
        self.button_select_file.setToolTip(self.button_select_file.objectName())

        # label_path_selected_file
        self.label_path_selected_file = PyQt5.QtWidgets.QLabel(self)
        self.label_path_selected_file.setObjectName('label_path_selected_file')
        self.label_path_selected_file.setEnabled(False)
        self.label_path_selected_file.setText(self.text_empty_path_file)
        self.label_path_selected_file.setGeometry(PyQt5.QtCore.QRect(10, 70, 400, 40))
        font = PyQt5.QtGui.QFont()
        font.setPointSize(10)
        self.label_path_selected_file.setFont(font)
        self.label_path_selected_file.adjustSize()
        self.label_path_selected_file.setToolTip(self.label_path_selected_file.objectName())

        #
        # тут место табличному виджету
        #

        # button_report_to_xls
        self.button_report_to_xls = PyQt5.QtWidgets.QPushButton(self)
        self.button_report_to_xls.setObjectName('button_report_to_xls')
        self.button_report_to_xls.setEnabled(False)
        self.button_report_to_xls.setText('Создать отчёт "Остатки на складе"')
        self.button_report_to_xls.setGeometry(PyQt5.QtCore.QRect(10, 230, 260, 25))
        self.button_report_to_xls.clicked.connect(self.report_to_xls)
        self.button_report_to_xls.setToolTip(self.button_report_to_xls.objectName())

        # button_exit
        self.button_exit = PyQt5.QtWidgets.QPushButton(self)
        self.button_exit.setObjectName('button_exit')
        self.button_exit.setText('Выход')
        self.button_exit.setGeometry(PyQt5.QtCore.QRect(10, 300, 100, 25))
        # self.button_exit.setFixedWidth(50)
        self.button_exit.clicked.connect(self.click_on_btn_exit)
        self.button_exit.setToolTip(self.button_exit.objectName())

    # нажатие на кнопку выбора файла
    def select_file(self):
        # переменная для хранения информации из окна выбора файла
        data_of_open_file_name = None

        # запоминание старого значения пути выбора файлов
        old_path_of_selected_file = self.label_path_selected_file.text()

        # окно выбора файла и переменная для хранения пути файла
        data_of_open_file_name = PyQt5.QtWidgets.QFileDialog.getOpenFileName(self,
                                                                             self.info_for_open_file,
                                                                             self.info_path_open_file,
                                                                             self.info_extention_open_file)
        # выбираю только путь из data_of_open_file_name
        selected_file_full_path = data_of_open_file_name[0]

        # нажата или не нажата кнопка выбора файла
        if selected_file_full_path == '':
            # если не выбран файл
            old_path_of_selected_file = self.label_path_selected_file.text()
            if self.label_path_selected_file.text() == self.text_empty_path_file:
                self.selected_file = None
            else:
                self.selected_file = old_path_of_selected_file
            self.label_path_selected_file.setText(old_path_of_selected_file)
            self.label_path_selected_file.adjustSize()
        else:
            # если выбран файл
            old_path_of_selected_file = self.label_path_selected_file.text()
            self.selected_file = os.path.normcase(selected_file_full_path)
            self.label_path_selected_file.setText(selected_file_full_path)
            self.label_path_selected_file.adjustSize()

        # определение расширения файла и выбор действий
        file_set = self.parse_file_parts(self.selected_file)

        if file_set:
            if file_set[3] == 'zip':
                self.take_zip(file_set)
            elif file_set[3] == 'csv':
                # эта функция для обработки данных для табличной части
                self.take_csv(file_set)
            else:
                # не могу определить расширение файла
                pass

    # активация объектов на форме зависящих от выбора файла
    def activate_objects(self):
        # TODO
        # это переписать

        # активация объектов на форме зависящих от выбора файла
        if self.text_empty_path_file not in self.label_path_selected_file.text():
            self.button_report_to_xls.setEnabled(True)

    # чтение zip файла
    def take_zip(self, take_file):
        # проверка на формат файла zip
        if not zipfile.is_zipfile(take_file[0]):
            # информационное окно про неправильный zip файл
            PyQt5.QtWidgets.QMessageBox.information(self,
                                                    'Ошибка',
                                                    f'Файл\n\n"{take_file[0]}"\n\n не является архивом zip')
        else:
            # файл экземпляр ZipFile
            zf = zipfile.ZipFile(take_file[0])

            # поиск файла csv в архиве zip перебором всех элементов архива
            for file_in_zf in zf.infolist():
                # если не папка, то продолжить
                if not file_in_zf.is_dir():
                    # проверка расширения
                    if self.parse_file_parts(file_in_zf.filename)[3].lower() == 'csv':
                        # сравнение имён файлов между именем архива и именем csv файла, они должны почти совпадать
                        diff_ratio_file_names = difflib.SequenceMatcher(
                                                    None,
                                                    take_file[2],
                                                    self.parse_file_parts(file_in_zf.filename)[2]).ratio()
                        if diff_ratio_file_names < 0.9:
                            PyQt5.QtWidgets.QMessageBox.information(
                                self,
                                'Ошибка',
                                f'Имя файла zip не совпадает с именам внутри архива.\n\n'
                                f'Выберите не переименованный файл скачанный с сайта\n\n'
                                f'или выберите другой.')
                        else:
                            # нужный файл найден, чтение из файла данных в бинарном виде
                            with zf.open(file_in_zf.filename) as file_csv_in_zip:
                                text = file_csv_in_zip.read()

                            # бинарные данные записываются во временный файл
                            with tempfile.NamedTemporaryFile(prefix='_from_zip_', suffix='.csv', delete=False) as fp:
                                fp.write(text)
                                # чтение переводится в начало файла
                                fp.seek(0)
                                # пока файл не закрыт получаю его имя
                                file_set = self.parse_file_parts(fp.name)
                                # и передаю его в обработку данных (первое это табличная часть, второе в эксель)
                                self.take_csv(file_set)

                            # закрываю файл
                            fp.close()
                    # else:
                    #     PyQt5.QtWidgets.QMessageBox.information(self, 'Ошибка',
                    #         f'Файл\n\n"{take_file[0]}"\n\n не содержит csv файла.')
            zf.close()

        # активация объектов на форме зависящих от выбора файла
        self.activate_objects()

    # чтение csv файла
    # и подготовка данных для выгрузки в табличную часть
    def take_csv(self, file_kit):
        # определение кодировки входных данных
        code_page = self.get_codepage(file_kit[0])

        # список для читаемых данных
        gathering_list = []
        # чтение данных из csv фйла построчно
        with open(file_kit[0], encoding=code_page, newline='') as fp:
            reader = csv.reader(fp)
            for key, row in enumerate(reader, start=1):
                # проверка на наличие в файле всех требующихся полей, поиск ведётся в первой строке
                if key == 1:
                    if not all(val in row for val in self.headers):
                        # информационное окно об отсутствии минимально необходимых полей в файле
                        PyQt5.QtWidgets.QMessageBox.information(
                            self,
                            'Ошибка',
                            f'Файл "{file_kit[1]}"\n\nне содержит всех нужных полей.\n\n'
                            f'Переформируйте файл с нужными или со всеми полями.\n\n'
                            f'{self.headers}')
                        break
                gathering_list.append(row)

            # если формируется не пустой список данных из csv, то передать его в подготовку для табличной части
            if len(gathering_list) > 1:
                # передача данных в подготовку для табличной части
                self.create_csv_list(gathering_list)

                # активация объектов на форме зависящих от выбора файла
                self.activate_objects()
            else:
                # информационное окно о пустом файле csv
                PyQt5.QtWidgets.QMessageBox.information(self, 'Ошибка', f'Файл пуст. Переформируйте файл.')

        # TODO
        # тут дальше дожна быть обработка и выгрузка в табличную часть,
        # но её я пока не написал

    # функция создания отчёта xls
    # данные обработать и вставить их в эксель на два листа
    def report_to_xls(self):
        # создаётся файл под отчёт, если данные имеются
        self.create_xls()

        # определение расширения файла и выбор действий
        file_set = self.parse_file_parts(self.selected_file)

        # определение кодировки входных данных
        code_page = self.get_codepage(file_set[0])

        headers = self.headers

        try:
            # прочитать весь файл
            df_all = pd.read_csv(file_set[0], encoding=code_page, dtype=object)

            # выбрать нужные колонки
            df = df_all[headers]
            # print(df.to_string())

            # посчитать количество prod_name
            q_prod_name = df.pivot_table('full_prod_name', 'prod_name', aggfunc='count', fill_value=0)
            q_prod_name.to_excel('output1.xlsx')
            # print(q_prod_name.to_string())

            # print()
            # подсчёт full_prod_name в колонке относительно prod_name
            df_group1 = df.pivot_table(['prod_name'], ['prod_name', 'full_prod_name', 'status', 'sgtin'],
                                       aggfunc='count', fill_value=0)
            df_group1.to_excel('output2.xlsx')
            # print(df_group1.to_string())

            df_group1 = df_group1.reset_index()
            # for index, row in df_group1.iterrows():
            #     for val in headers:
            #         print(f'{row[val] = }')
            #     print('*' * 155)
            df_group1.to_excel('output3.xlsx')
        except pd.errors.EmptyDataError:
            pass
        except NameError:
            pass
        except KeyError:
            pass
        except pd.errors.ParserError:
            pass

    # функция создания списка с данными по шаблону self.headers, чтобы колонки шли в порядке self.headers
    def create_csv_list(self, input_list: list):
        # нахождение мест заголовков во входящих данных, на случай если колонки во входящих данных перепутаны
        place_dict = {}
        for header in self.headers:
            place = input_list[0].index(header)
            place_dict[header] = place

        # формирование выходного списка согласно порядку заголовков
        output_list = []
        for key, list_of_input_list in enumerate(input_list):
            str_list = []
            for place in place_dict:
                str_list.append(list_of_input_list[place_dict[place]])
            output_list.append(str_list)

        # TODO
        # тут вызвать выгрузку в табличную часть или это сделать в take_csv (252 строка)
        # print(*output_list, sep='\n')
        return output_list

    # функция создания файла xls для отчёта
    @staticmethod
    def create_xls():
        # создание книги xls и двух листов, для общего отчёта и детального
        wb = openpyxl.Workbook()
        wb_s = wb.active
        wb_s.title = 'Общий'
        wb.create_sheet('Структурный')
        wb.create_sheet('Детальный')
        wb.save('out.xlsx')

    # получение кодировки файла
    @staticmethod
    def get_codepage(one_file):
        detector = chardet.universaldetector.UniversalDetector()
        with open(one_file, 'rb') as fh:
            for line in fh:
                detector.feed(line)
                if detector.done:
                    break
            detector.close()
        # print(detector.result)
        return detector.result['encoding']

    # функция разбора файла на полный путь, полное имя файла, имя файла и расширение файла
    @staticmethod
    def parse_file_parts(take_file):
        # проверка на непустой путь, то есть выбранный файл
        if take_file:
            file_full_path = take_file
            file_full_name = os.path.basename(take_file)
            file_name = file_full_name.rsplit('.', 1)[0]
            if len(file_full_name.rsplit('.', 1)) > 1:
                file_ext = file_full_name.rsplit('.', 1)[1]
            else:
                file_ext = ''

            return file_full_path, file_full_name, file_name, file_ext
        else:
            return None

    # событие - нажатие на кнопку Выход
    @staticmethod
    def click_on_btn_exit():
        sys.exit()


# создание основного окна
def main_app():
    app = PyQt5.QtWidgets.QApplication(sys.argv)
    app.setStyle('Fusion')
    app_window_main = WindowMain()
    app_window_main.show()
    sys.exit(app.exec_())


# запуск основного окна
if __name__ == '__main__':
    main_app()
